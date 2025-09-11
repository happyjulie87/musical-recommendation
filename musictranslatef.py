from selenium import webdriver #瀏覽器
from selenium.webdriver.chrome.service import Service #抓檔案位置
from selenium.webdriver.common.by import By #html定位器
from selenium.webdriver.chrome.options import Options #設置瀏覽器設定
from selenium.webdriver.support.wait import WebDriverWait #離覽器等待
from selenium.webdriver.support import expected_conditions as EC #判斷目標可被互動
from selenium.webdriver.common.keys import Keys  #模擬輸入功能鍵
from selenium.webdriver.support.ui import Select #下拉選單
from time import sleep 
import pandas as pd  # 加在檔案最上方的 import 區
import openpyxl
import os

from deep_translator import GoogleTranslator # 匯入翻譯工具


# 建立一個 Chrome 瀏覽器的選項（Options）物件，讓你可以設定啟動 Chrome 時的各種參數。
option=webdriver.ChromeOptions()
# 建立一個設定，用來封鎖所有網站跳出來的『是否允許通知』的請求。
prefs={"profile.default_content_setting_values":{"notifications":2}}
# 套用prefs設定
option.add_experimental_option("prefs",prefs)
# 讓瀏覽器在腳本結束後保持開啟
option.add_experimental_option("detach", True)

driver=webdriver.Chrome(service=Service(executable_path="d:\\python_test\\chromedriver.exe"),options=option)

driver.get("https://kopis.or.kr/por/boxoffice/boxoffice.do?menuLD=mnu_00024")

#音樂的
search_btn1 = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="su_con"]/div[2]/div[1]/div/ul/li[3]/p/button/span')))
search_btn1.click()

#每年的
search_btn2 = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="listForm"]/div/div[2]/div/button[4]')))
search_btn2.click()

all_data = []
all_data.append(["排名", "標題", "日期", "地點", "圖片連結", "標題_韓文", "地點_韓文","連結"])

DATE1_elem = driver.find_element(By.XPATH,'//*[@id="clsmonth"]/div/div[1]/div/div/div/div/input')

orig_date = DATE1_elem.get_attribute('value')
# 字串分隔 年.月.日
parts = orig_date.split('.')

new_year = str(int(parts[0]))
print(new_year)

for i in range(0, 5):
     
    new_year = str(int(parts[0]) - i)
    print(new_year)

    DATE1= driver.find_element(By.CLASS_NAME,"DatePicker_btnDatepicker__edI76")
    #JS 點擊，不能用 click()，會被擋住
    driver.execute_script("arguments[0].click();", DATE1)


    YEAR1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="clsmonth"]/div/div[1]/div/div[2]/div[2]/div/div/div/div/div[1]/div[1]/div/select[1]')))
    Select(YEAR1).select_by_value(new_year)

    MON1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="clsmonth"]/div/div[1]/div/div[2]/div[2]/div/div/div/div/div[1]/div[1]/div/select[2]')))
    Select(MON1).select_by_value("0")

    DAY1 = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="clsmonth"]//div[text()="1"]')))
    driver.execute_script("arguments[0].click();", DAY1)

    search_btn3 = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="listForm"]/div/div[4]/button')))
    search_btn3.click()

    sleep(20)    

    # 等待資料載入完成
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'li > div > div > span.Poster_num__Ez8Q1'))
    )



    lis = driver.find_elements(By.CSS_SELECTOR, 'li > div > div > span.Poster_num__Ez8Q1')
    for span in lis:
        li = span.find_element(By.XPATH, './ancestor::li')
        # 排名
        rank = str(new_year+"-"+span.text)
        # 標題
        title = li.find_element(By.CSS_SELECTOR, 'div.Poster_zi__4g1td h4 a').text
        # 日期
        date = li.find_element(By.CSS_SELECTOR, 'div.Poster_zi__4g1td div p:nth-child(1)').text
        # 地點
        place = li.find_element(By.CSS_SELECTOR, 'div.Poster_zi__4g1td div p:nth-child(2)').text
        # 圖片連結
        img = li.find_element(By.CSS_SELECTOR, 'div.Poster_tu__k7uUQ img').get_attribute('src')

        title_kr = li.find_element(By.CSS_SELECTOR, 'div.Poster_zi__4g1td h4 a').text

        place_kr = li.find_element(By.CSS_SELECTOR, 'div.Poster_zi__4g1td div p:nth-child(2)').text

        link = li.find_element(By.CSS_SELECTOR, 'div.Poster_zi__4g1td h4 a').get_attribute('href')

        all_data.append([rank, title, date, place, img, title_kr, place_kr,link])
    

# # parts = orig_date.split('.')
# # new_year = str(int(parts[0]) - 1)
# # new_date = new_year + '.' + parts[1] + '.' + parts[2]
# # # 清空並輸入新日期






wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
wb.save('musicitemf.xlsx') 

wb = openpyxl.load_workbook('musicitemf.xlsx', data_only=True)  

s3 = wb.create_sheet('musiclist')     # 新增工作表

# 建立一個翻譯工具的物件
translator = GoogleTranslator(source='ko', target='zh-TW')

# 取得所有需要翻譯的標題和地點
titles_to_translate = [row[1] for row in all_data[1:]]
places_to_translate = [row[3] for row in all_data[1:]]

# 一次性翻譯所有標題和地點
translated_titles = translator.translate_batch(titles_to_translate)
translated_places = translator.translate_batch(places_to_translate)

# 將翻譯結果更新回 all_data
for i, row in enumerate(all_data[1:]):
    row[1] = translated_titles[i]
    row[3] = translated_places[i]

for i in all_data:
    s3.append(i)                   # 逐筆添加到最後一列

if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

wb.save('musicitemf.xlsx')

driver.quit()
